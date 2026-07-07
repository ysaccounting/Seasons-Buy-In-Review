"""
Season Ticket Buy-In Review — backend.

Reconciles the HAL-NFL season-ticket database against the TicketVault
Purchase Details exports. Every HAL record (one seat block — tickets or
parking) is checked against what is actually loaded in TicketVault.

A HAL record is RECONCILED when both are true:
  * the TicketVault total cost for that seat block ties to the HAL total
    cost, within a tolerance (default $1.00), and
  * the number of TicketVault games carrying a cost equals the HAL # games.

Otherwise it is NOT RECONCILED, with a note:
  * "Not bought in"                              — no matching rows in TicketVault
  * "total cost not equal"                       — present, cost doesn't tie
  * "# games not equal"                          — present, games-with-cost differ
  * "total cost not equal, # games not equal"    — both

Match key: email + team + section/row + individual seat number. Direction is
HAL -> TicketVault only (vault rows with no HAL record are ignored).

Multiple companies can be uploaded together. Records are split by a "Company"
column in the HAL file (falling back to the file name), and EACH COMPANY gets
its own workbook. League and Year are chosen in the UI and drive the file name:

    Seasons Review - {Company} - {League} - {Year}.xlsx

Output workbook tabs: Summary (with Company / League / Year), Reconciled,
Not Reconciled.
"""

import io
import os
import csv
import re
import time
import uuid
import zipfile
import tempfile
import datetime as dt
from collections import defaultdict, OrderedDict

from flask import Flask, request, jsonify, send_file, send_from_directory, abort
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder=None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STORE_DIR = os.path.join(tempfile.gettempdir(), "recon_store")
os.makedirs(STORE_DIR, exist_ok=True)

MAPPING_FILE = os.path.join(BASE_DIR, "Master_Mapping_List.xlsx")
EXCLUDE_COMPANIES = {"needle", "damona"}


def load_company_names():
    """Company/broker Short Names for the per-file dropdown, from the Master
    Mapping List (col 'Short Name'). Falls back to a built-in list."""
    try:
        wb = load_workbook(MAPPING_FILE, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [str(c or "").strip().lower() for c in rows[0]]
        ci = header.index("short name") if "short name" in header else 1
        names, seen = [], set()
        for r in rows[1:]:
            v = str(r[ci]).strip() if ci < len(r) and r[ci] else ""
            if v and v.lower() not in EXCLUDE_COMPANIES and v not in seen:
                seen.add(v)
                names.append(v)
        if names:
            return names
    except Exception:
        pass
    return ["TTG", "Y&S", "YourTix", "Asher", "Chase", "Katz", "Levine",
            "Levovitz", "TL", "GK", "Grossman", "Pollak", "Sternbuch", "Waxler"]


COMPANY_NAMES = load_company_names()

# Preferred display order for the company dropdown (mapping-list membership still
# decides which companies exist; this only orders them).
PREFERRED_COMPANY_ORDER = ["Y&S", "Grossman", "Sternbuch", "Pollak", "Levine",
                           "Levovitz", "Chase", "Asher", "Katz", "GK", "TL",
                           "Waxler", "TTG", "YourTix"]
COMPANY_NAMES = ([c for c in PREFERRED_COMPANY_ORDER if c in COMPANY_NAMES]
                 + [c for c in COMPANY_NAMES if c not in PREFERRED_COMPANY_ORDER])

# Secondary-market vendors to exclude from Purchase Details before reconciling.
EXCLUDED_VENDORS = ("ticketmaster", "tickpick", "stubhub", "ticket evolution", "gotickets")


def load_broker_pvcompanies():
    """From the Master Mapping List, map each broker Short Name to the set of
    TicketVault 'Company' values that belong to it (col 'TicketVault Company /
    Applied Payments Category'). Returns (broker->set(pv_company_lower),
    pv_company_lower->broker)."""
    broker_to_pv, pv_to_broker = {}, {}
    try:
        wb = load_workbook(MAPPING_FILE, data_only=True, read_only=True)
        rows = list(wb.worksheets[0].iter_rows(values_only=True))
        wb.close()
        header = [str(c or "").strip().lower() for c in rows[0]]
        si = header.index("short name") if "short name" in header else 1
        ti = next((i for i, h in enumerate(header) if h.startswith("ticketvault company")), 3)
        for r in rows[1:]:
            short = str(r[si]).strip() if si < len(r) and r[si] else ""
            pv = str(r[ti]).strip().lower() if ti < len(r) and r[ti] else ""
            if not short or not pv or short.lower() in EXCLUDE_COMPANIES:
                continue
            broker_to_pv.setdefault(short, set()).add(pv)
            pv_to_broker[pv] = short
    except Exception:
        pass
    return broker_to_pv, pv_to_broker


BROKER_PVCOMPANIES, PVCOMPANY_TO_BROKER = load_broker_pvcompanies()


def _is_excluded_vendor(vendor):
    v = str(vendor or "").strip().lower()
    return any(v.startswith(x) for x in EXCLUDED_VENDORS)

# --------------------------------------------------------------------------- #
# Team-name normalization
#
# Brokers write team names every which way — "Bengals", "CIN", "49ers", full
# "Cincinnati Bengals". TicketVault uses the full "City Nickname". We map both
# sides to that canonical form, scoped by the League chosen in the UI (so
# "Cardinals" -> Arizona under NFL). Nicknames and single-team cities are
# derived automatically from each league's team list; abbreviations and
# historical/edge aliases are listed explicitly. Leagues without a list here
# pass through unchanged.
# --------------------------------------------------------------------------- #

NFL_TEAMS = [
    "Arizona Cardinals", "Atlanta Falcons", "Baltimore Ravens", "Buffalo Bills",
    "Carolina Panthers", "Chicago Bears", "Cincinnati Bengals", "Cleveland Browns",
    "Dallas Cowboys", "Denver Broncos", "Detroit Lions", "Green Bay Packers",
    "Houston Texans", "Indianapolis Colts", "Jacksonville Jaguars", "Kansas City Chiefs",
    "Las Vegas Raiders", "Los Angeles Chargers", "Los Angeles Rams", "Miami Dolphins",
    "Minnesota Vikings", "New England Patriots", "New Orleans Saints", "New York Giants",
    "New York Jets", "Philadelphia Eagles", "Pittsburgh Steelers", "San Francisco 49ers",
    "Seattle Seahawks", "Tampa Bay Buccaneers", "Tennessee Titans", "Washington Commanders",
]
NFL_ALIASES = {
    "ari": "Arizona Cardinals", "arz": "Arizona Cardinals", "atl": "Atlanta Falcons",
    "bal": "Baltimore Ravens", "blt": "Baltimore Ravens", "buf": "Buffalo Bills",
    "car": "Carolina Panthers", "chi": "Chicago Bears", "cin": "Cincinnati Bengals",
    "cle": "Cleveland Browns", "dal": "Dallas Cowboys", "den": "Denver Broncos",
    "det": "Detroit Lions", "gb": "Green Bay Packers", "gnb": "Green Bay Packers",
    "hou": "Houston Texans", "ind": "Indianapolis Colts", "jax": "Jacksonville Jaguars",
    "jac": "Jacksonville Jaguars", "jags": "Jacksonville Jaguars", "kc": "Kansas City Chiefs",
    "kan": "Kansas City Chiefs", "lv": "Las Vegas Raiders", "lvr": "Las Vegas Raiders",
    "oak": "Las Vegas Raiders", "oakland": "Las Vegas Raiders", "raiders": "Las Vegas Raiders",
    "lac": "Los Angeles Chargers", "sd": "Los Angeles Chargers", "san diego": "Los Angeles Chargers",
    "lar": "Los Angeles Rams", "st louis rams": "Los Angeles Rams", "mia": "Miami Dolphins",
    "min": "Minnesota Vikings", "ne": "New England Patriots", "nwe": "New England Patriots",
    "pats": "New England Patriots", "no": "New Orleans Saints", "nor": "New Orleans Saints",
    "nyg": "New York Giants", "nyj": "New York Jets", "phi": "Philadelphia Eagles",
    "pit": "Pittsburgh Steelers", "sf": "San Francisco 49ers", "sfo": "San Francisco 49ers",
    "niners": "San Francisco 49ers", "sea": "Seattle Seahawks", "tb": "Tampa Bay Buccaneers",
    "tbb": "Tampa Bay Buccaneers", "bucs": "Tampa Bay Buccaneers", "ten": "Tennessee Titans",
    "was": "Washington Commanders", "wsh": "Washington Commanders",
    "washington football team": "Washington Commanders", "redskins": "Washington Commanders",
}
MLB_TEAMS = [
    "Arizona Diamondbacks", "Atlanta Braves", "Baltimore Orioles", "Boston Red Sox",
    "Chicago Cubs", "Chicago White Sox", "Cincinnati Reds", "Cleveland Guardians",
    "Colorado Rockies", "Detroit Tigers", "Houston Astros", "Kansas City Royals",
    "Los Angeles Angels", "Los Angeles Dodgers", "Miami Marlins", "Milwaukee Brewers",
    "Minnesota Twins", "New York Mets", "New York Yankees", "Oakland Athletics",
    "Philadelphia Phillies", "Pittsburgh Pirates", "San Diego Padres", "San Francisco Giants",
    "Seattle Mariners", "St. Louis Cardinals", "Tampa Bay Rays", "Texas Rangers",
    "Toronto Blue Jays", "Washington Nationals",
]
NBA_TEAMS = [
    "Atlanta Hawks", "Boston Celtics", "Brooklyn Nets", "Charlotte Hornets",
    "Chicago Bulls", "Cleveland Cavaliers", "Dallas Mavericks", "Denver Nuggets",
    "Detroit Pistons", "Golden State Warriors", "Houston Rockets", "Indiana Pacers",
    "Los Angeles Clippers", "Los Angeles Lakers", "Memphis Grizzlies", "Miami Heat",
    "Milwaukee Bucks", "Minnesota Timberwolves", "New Orleans Pelicans", "New York Knicks",
    "Oklahoma City Thunder", "Orlando Magic", "Philadelphia 76ers", "Phoenix Suns",
    "Portland Trail Blazers", "Sacramento Kings", "San Antonio Spurs", "Toronto Raptors",
    "Utah Jazz", "Washington Wizards",
]
NHL_TEAMS = [
    "Anaheim Ducks", "Boston Bruins", "Buffalo Sabres", "Calgary Flames",
    "Carolina Hurricanes", "Chicago Blackhawks", "Colorado Avalanche", "Columbus Blue Jackets",
    "Dallas Stars", "Detroit Red Wings", "Edmonton Oilers", "Florida Panthers",
    "Los Angeles Kings", "Minnesota Wild", "Montreal Canadiens", "Nashville Predators",
    "New Jersey Devils", "New York Islanders", "New York Rangers", "Ottawa Senators",
    "Philadelphia Flyers", "Pittsburgh Penguins", "San Jose Sharks", "Seattle Kraken",
    "St. Louis Blues", "Tampa Bay Lightning", "Toronto Maple Leafs", "Utah Hockey Club",
    "Vancouver Canucks", "Vegas Golden Knights", "Washington Capitals", "Winnipeg Jets",
]


def _build_team_index(full_names, extra_aliases):
    """canonical lookup: alias(lowercased) -> full team name. Nicknames and
    unambiguous cities are derived from the team list automatically."""
    alias, nick, city = {}, {}, {}
    for full in full_names:
        parts = full.split()
        alias[full.lower()] = full
        nick.setdefault(parts[-1].lower(), []).append(full)
        city.setdefault(" ".join(parts[:-1]).lower(), []).append(full)
    for key, teams in nick.items():
        if len(teams) == 1:
            alias[key] = teams[0]
    for key, teams in city.items():
        if key and len(teams) == 1:
            alias[key] = teams[0]
    for a, full in extra_aliases.items():   # explicit wins over derived
        alias[a.lower()] = full
    return alias


TEAM_INDEX = {
    "NFL": _build_team_index(NFL_TEAMS, NFL_ALIASES),
    "MLB": _build_team_index(MLB_TEAMS, {}),
    "NBA": _build_team_index(NBA_TEAMS, {}),
    "NHL": _build_team_index(NHL_TEAMS, {}),
}


def _normalize_team(name, league):
    raw = re.sub(r"\s+", " ", str(name or "").strip())
    if not raw:
        return ""
    idx = TEAM_INDEX.get(league, {})
    # try the name as-is, then progressively strip trailing qualifiers like
    # " - New", " (New)", " - Renewal" — accepting a stripped form only if it
    # resolves to a real team, so we never mangle a genuine name.
    candidates, s = [raw.lower()], raw
    for _ in range(3):
        m = re.sub(r"\s*(?:-\s*[^-()]+|\([^()]*\))\s*$", "", s).strip()
        if m == s or not m:
            break
        candidates.append(m.lower())
        s = m
    for cand in candidates:
        if cand in idx:
            return idx[cand]
    return raw

DEFAULT_TOLERANCE = 1.00  # dollars; TV total must tie to HAL total within this
LEAGUES = ["MLB", "MLS", "NBA", "NFL", "NHL", "NCAAF", "NCAAB", "WNBA", "Racing"]
YEARS = ["2025-26", "2026-27", "2027-28", "2028-29"]

# Different companies label their HAL columns differently. Each HAL field maps to
# every header name we've seen for it (case-insensitive). To support a new company
# whose layout differs, add its header names to the relevant list here.
HAL_SYNONYMS = {
    "company": ["Company", "Client", "Broker", "Group"],
    "email":   ["Email", "Account Email", "Email Address", "Login Email",
                "Email (from Profiles)", "Account EMAIL", "user Name/email",
                "Username/Email", "User Name/Email", "PO Email Account"],
    "team":    ["Team", "Team Name"],
    "fp":      ["Full/Partial", "Type", "Plan", "Plan Type", "Ticket Type",
                "Package", "Plan/Parking"],
    "section": ["Section", "Sec", "Sec.", "Sction"],
    "row":     ["Row"],
    "seats":   ["Seats", "Seat", "Seat(s)", "Seat Range", "Seat #", "Seat Numbers"],
    "qty":     ["Qty", "Quantity", "# Seats", "Seat Qty", "# of Seats"],
    "games":   ["Games/Threshold", "Games Threshold", "# Games", "Games", "Games#",
                "# of Games", "Num Games", "Game Count"],
    "total":   ["Total", "Total Cost", "Total Price", "Total Amount", "Cost",
                "Amount Paid", "Amount $", "Amount ($)", "Amount"],
}

# --------------------------------------------------------------------------- #
# Generic helpers
# --------------------------------------------------------------------------- #

_NUM_FORMULA = re.compile(r"-?\d+(\.\d+)?")


def _cleanup_old(max_age_seconds=12 * 3600):
    now = time.time()
    for name in os.listdir(STORE_DIR):
        path = os.path.join(STORE_DIR, name)
        try:
            if now - os.path.getmtime(path) > max_age_seconds:
                if os.path.isdir(path):
                    for f in os.listdir(path):
                        os.remove(os.path.join(path, f))
                    os.rmdir(path)
                else:
                    os.remove(path)
        except OSError:
            pass


def _amount(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", " "):
        return None
    if s.startswith("="):
        body = s[1:]
        return float(body) if _NUM_FORMULA.fullmatch(body) else None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", "."):
        return None
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None


def _rows_from_upload(filename, data):
    low = filename.lower()
    if low.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return [list(r) for r in csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    best, best_score = None, -1
    for ws in wb.worksheets:
        score = (ws.max_row or 0) * (ws.max_column or 1)
        if score > best_score:
            best, best_score = ws, score
    rows = [list(r) for r in best.iter_rows(values_only=True)]
    wb.close()
    return rows


def _norm_header(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _find_header_row(rows):
    for i, row in enumerate(rows[:15]):
        cells = [_norm_header(c) for c in row]
        if len([c for c in cells if c]) >= 2:
            return i, cells
    return 0, [_norm_header(c) for c in (rows[0] if rows else [])]


def _col_index(header, *names):
    wanted = [_norm_header(n) for n in names]
    for i, h in enumerate(header):
        if h in wanted:
            return i
    return None


def _cell(row, i):
    return row[i] if (i is not None and i < len(row)) else None


def _safe_name(s):
    """Sanitize a string for use in a file name."""
    return re.sub(r'[\\/:*?"<>|]+', " ", str(s or "")).strip() or "Unknown"


def _company_from_filename(filename):
    stem = os.path.splitext(os.path.basename(filename))[0]
    return stem.strip() or "Unknown"


def _fmt_asof(as_of):
    """Format the date picker's YYYY-MM-DD as m-d-yy (e.g. 2026-07-07 -> 7-7-26)."""
    try:
        d = dt.datetime.strptime(str(as_of).strip(), "%Y-%m-%d")
        return f"{d.month}-{d.day}-{d.strftime('%y')}"
    except ValueError:
        return str(as_of).strip()


# --------------------------------------------------------------------------- #
# Seat / key normalization
# --------------------------------------------------------------------------- #

def _emails(cell):
    return [x.strip().lower() for x in re.split(r"[,;]", str(cell or ""))
            if x.strip() and x.strip().lower() != "nan"]


def _team(cell):
    return str(cell or "").strip()


def _sec(cell):
    return re.sub(r"\s+", " ", str(cell or "").strip().upper())


def _row(cell):
    return str(cell or "").strip().upper()


def _seatnums(cell):
    # Excel silently turns seat ranges like "3-9" into a date (Mar 9). Recover
    # them as month-day -> seat range (verified against Qty in real exports).
    if isinstance(cell, (dt.datetime, dt.date)):
        lo, hi = sorted((cell.month, cell.day))
        return set(range(lo, hi + 1))
    s = str(cell or "").replace(" ", "")
    m = re.match(r"^(\d+)-(\d+)$", s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if a <= b and b - a < 60:
            return set(range(a, b + 1))
    if re.match(r"^\d+$", s):
        return {int(s)}
    return set()


def _seat_text(cell):
    """Readable seat string. Excel turns ranges like '3-9' into a date (Mar 9);
    render those back as 'month-day' so both display and matching are correct."""
    if isinstance(cell, (dt.datetime, dt.date)):
        return f"{cell.month}-{cell.day}"
    return str(cell or "").strip()


_PARK_RE = re.compile(r"\b(parking|garage|lot|prkg|pkg)\b", re.I)


def _looks_parking(*vals):
    return any(_PARK_RE.search(str(v or "")) for v in vals)


def _is_parking_hal(full_partial, section="", row=""):
    # Type/Plan says parking, or the section/row names a lot/garage.
    return "parking" in str(full_partial or "").lower() or _looks_parking(section, row)


def _is_parking_pv(sec, row, venue=""):
    # Sec/row names a lot, or the TicketVault Venue is a parking venue.
    return _looks_parking(sec, row, venue)


# Plan/Type values that mean the seats aren't actually held this season and so
# shouldn't be reconciled. (You may also strip these before uploading.)
NONACTIVE_KEYWORDS = ("deposit", "wait list", "waitlist", "waitlisted", "inquir",
                      "notification", "cancel", "nothing for this", "back and forth",
                      "pending", "declined", "not renew")


def _is_nonactive(full_partial):
    s = str(full_partial or "").lower()
    return any(k in s for k in NONACTIVE_KEYWORDS)


_EMAIL_RE = re.compile(r"[^@\s,;]+@[^@\s,;]+\.[^@\s,;]+")


def _year_tokens(year):
    m = re.match(r"^(\d{4})-(\d{2})$", str(year or "").strip())
    return (m.group(1), m.group(1)[2:], m.group(2)) if m else None


def _year_total_cols(year):
    """Some HALs keep total cost in a season-year column: Levovitz '26/27',
    GK '26 Total'. Build those names from the selected year."""
    t = _year_tokens(year)
    if not t:
        return []
    yyyy, yy, nn = t
    return [f"{yy} Total", f"{yyyy} Total", f"{yy}/{nn}", f"{yy}-{nn}"]


def _year_plan_cols(year):
    """Some HALs name the plan/type column by year (GK '2026 Plan')."""
    t = _year_tokens(year)
    if not t:
        return []
    yyyy, yy, _nn = t
    return [f"{yyyy} Plan", f"{yy} Plan"]


def _sniff_email_col(header, data_rows):
    """When no email header matches (GK stores it under 'Profiles', TL under
    'Name'), pick the column whose values most look like email addresses."""
    best_i, best_frac = None, 0.0
    sample = data_rows[:200]
    for i in range(len(header)):
        vals = [r[i] for r in sample if i < len(r) and r[i] not in (None, "")]
        if not vals:
            continue
        frac = sum(1 for v in vals if _EMAIL_RE.search(str(v))) / len(vals)
        if frac > best_frac:
            best_i, best_frac = i, frac
    return best_i if best_frac >= 0.5 else None


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #

def parse_hal(rows, filename, company, year="", league=""):
    """Parse a HAL season-ticket database. One record per seat block. Column
    names vary by company and are matched via HAL_SYNONYMS (email is also
    sniffed from values). Team names are normalized to TicketVault's canonical
    form for the chosen league. Non-active rows (deposits, waitlist, inquiries)
    are dropped. Returns (records, excluded_count)."""
    hidx, header = _find_header_row(rows)
    ci = {field: _col_index(header, *names) for field, names in HAL_SYNONYMS.items()}
    data_rows = rows[hidx + 1:]
    if ci["email"] is None:
        ci["email"] = _sniff_email_col(header, data_rows)
    if ci["email"] is None or ci["team"] is None:
        raise ValueError(
            f"{os.path.basename(filename)}: couldn't find an email or team column. "
            f"Looked for email as {HAL_SYNONYMS['email']} (and by sniffing values) "
            f"and team as {HAL_SYNONYMS['team']}. Add this file's header names to "
            f"HAL_SYNONYMS.")
    if ci["total"] is None:
        ci["total"] = _col_index(header, *_year_total_cols(year))
    if ci["fp"] is None:
        ci["fp"] = _col_index(header, *_year_plan_cols(year))
    out, excluded = [], 0
    for row in data_rows:
        team = _normalize_team(_cell(row, ci["team"]), league)
        emails = _emails(_cell(row, ci["email"]))
        if not team or not emails:
            continue
        fp = str(_cell(row, ci["fp"]) or "").strip()
        if _is_nonactive(fp):
            excluded += 1
            continue
        games = _amount(_cell(row, ci["games"]))
        out.append({
            "company": company,
            "emails": emails,
            "team": team,
            "Full/Partial": fp,
            "Section": str(_cell(row, ci["section"]) or "").strip(),
            "Row": str(_cell(row, ci["row"]) or "").strip(),
            "Seats": _seat_text(_cell(row, ci["seats"])),
            "Qty": str(_cell(row, ci["qty"]) or "").strip(),
            "Email": str(_cell(row, ci["email"]) or "").strip(),
            "games": int(games) if games is not None else None,
            "total": _amount(_cell(row, ci["total"])) or 0.0,
            "sec_n": _sec(_cell(row, ci["section"])),
            "row_n": _row(_cell(row, ci["row"])),
            "is_parking": _is_parking_hal(fp, _cell(row, ci["section"]), _cell(row, ci["row"])),
        })
    return out, excluded


def parse_details(rows, filename, league=""):
    """Parse a TicketVault Purchase Details export. Returns
    (parsed_rows, header, usable) where each parsed row carries its normalized
    team, seat data, company (for scoping), original row (for the source tab),
    and an `excluded_vendor` flag. Secondary-market vendor rows are kept for the
    source tab but flagged so they're skipped during matching."""
    hidx, header = _find_header_row(rows)
    raw_header = list(rows[hidx]) if hidx < len(rows) else []
    ci = {
        "email": _col_index(header, "PO Email Account", "Email"),
        "team": _col_index(header, "Team/Performer", "Team"),
        "sec": _col_index(header, "Sec", "Section"),
        "row": _col_index(header, "Row"),
        "seats": _col_index(header, "Seats"),
        "event": _col_index(header, "Event Date"),
        "cost": _col_index(header, "Total Cost"),
        "company": _col_index(header, "Company"),
        "vendor": _col_index(header, "Vendor"),
        "venue": _col_index(header, "Venue"),
    }
    if ci["email"] is None or ci["team"] is None or ci["cost"] is None:
        raise ValueError(f"{os.path.basename(filename)}: Purchase Details file is "
                         f"missing 'PO Email Account', 'Team/Performer' or 'Total Cost'.")
    parsed, usable = [], 0
    for row in rows[hidx + 1:]:
        email = str(_cell(row, ci["email"]) or "").strip().lower()
        team = _normalize_team(_cell(row, ci["team"]), league)
        if not email or not team:
            continue
        excluded_vendor = _is_excluded_vendor(_cell(row, ci["vendor"]))
        sec = _sec(_cell(row, ci["sec"]))
        rw = _row(_cell(row, ci["row"]))
        parsed.append({
            "email": email, "team": team, "sec": sec, "row": rw,
            "seatset": _seatnums(_cell(row, ci["seats"])),
            "event": str(_cell(row, ci["event"]) or ""),
            "cost": _amount(_cell(row, ci["cost"])) or 0.0,
            "is_parking": _is_parking_pv(sec, rw, _cell(row, ci["venue"])),
            "company_norm": str(_cell(row, ci["company"]) or "").strip().lower(),
            "excluded_vendor": excluded_vendor,   # kept for source, skipped for matching
            "raw": list(row),
        })
        if not excluded_vendor:
            usable += 1
    return parsed, raw_header, usable


# --------------------------------------------------------------------------- #
# Reconciliation
# --------------------------------------------------------------------------- #

def _games_split(matches):
    """Return (games_with_cost, games_without_cost, total_cost) for a set of
    matching vault rows. A game counts as 'with cost' if any matching row for
    that event carries a positive cost."""
    ev = defaultdict(float)
    total = 0.0
    for x in matches:
        ev[x["event"]] = max(ev[x["event"]], x["cost"])
        total += x["cost"]
    wc = sum(1 for v in ev.values() if v > 0)
    woc = sum(1 for v in ev.values() if v == 0)
    return wc, woc, round(total, 2)


def reconcile(hal_rows, primary_index, secondary_index, tolerance):
    reconciled, not_reconciled = [], []
    for r in hal_rows:
        own = set(r["emails"])
        vr = []
        for e in r["emails"]:
            vr.extend(primary_index.get((e, r["team"]), []))

        if not vr:
            own_matches = []
        elif r["is_parking"]:
            # try an exact lot/seat match first so multiple spots don't get
            # summed against each single HAL row; fall back to all of the team's
            # parking when the lot/seat numbers don't line up.
            hs = _seatnums(r["Seats"])
            seat_m = [x for x in vr if x["is_parking"] and x["sec"] == r["sec_n"]
                      and x["row"] == r["row_n"] and hs and (x["seatset"] & hs)]
            own_matches = seat_m if seat_m else [x for x in vr if x["is_parking"]]
        else:
            hs = _seatnums(r["Seats"])
            own_matches = [x for x in vr if (not x["is_parking"])
                           and x["sec"] == r["sec_n"] and x["row"] == r["row_n"]
                           and (x["seatset"] & hs)]

        wc, woc, tv_cost = _games_split(own_matches)
        # $0 parking: HAL carries no cost, so every game is expected at $0 in TV.
        # Compare HAL # games against TV games WITHOUT cost instead of with cost.
        zero_parking = r["is_parking"] and r["total"] == 0
        base = {
            "Team": r["team"], "Email": r["Email"], "Full/Partial": r["Full/Partial"],
            "Section": r["Section"], "Row": r["Row"], "Seats": r["Seats"], "Qty": r["Qty"],
            "# Games": r["games"], "HAL Total Cost": round(r["total"], 2),
            "TV Total Cost": tv_cost, "# Games w/Cost": wc, "# Games w/o Cost": woc,
            "_parking": r["is_parking"],
        }

        def variances(tvc, games_wc, games_woc, alt_email=""):
            dollar = round(tvc - r["total"], 2)
            pct = round(dollar / r["total"], 4) if r["total"] else None
            ref = games_woc if zero_parking else games_wc
            return {
                "Var Total Cost": dollar,
                "Var Total Cost %": pct,
                "Var # Games w/Cost": (ref - r["games"]) if r["games"] is not None else None,
                "Var Email Address": alt_email,
            }

        if own_matches:
            cost_ok = abs(tv_cost - r["total"]) <= tolerance
            games_known = r["games"] is not None
            if zero_parking:
                games_ok = (not games_known) or (woc >= r["games"])
            else:
                games_ok = games_known and wc == r["games"]
            if cost_ok and games_ok:
                reconciled.append(base)
                continue
            parts = []
            if not cost_ok:
                parts.append("total cost not equal")
            if not games_ok:
                parts.append("# games not in HAL"
                             if (not games_known and not zero_parking) else "# games not equal")
            not_reconciled.append({**base, **variances(tv_cost, wc, woc),
                                    "Notes": ", ".join(parts), "_p": 1})
            continue

        # nothing under the account's own email — look for the same seats under a
        # DIFFERENT email in this broker's vault (same team/sec/row, seat overlap)
        if not r["is_parking"]:
            hs = _seatnums(r["Seats"])
            cands = secondary_index.get((r["team"], r["sec_n"], r["row_n"]), [])
            by_email = defaultdict(list)
            for x in cands:
                if x["email"] not in own and (x["seatset"] & hs):
                    by_email[x["email"]].append(x)
            for alt_email, ms in by_email.items():
                a_wc, a_woc, a_cost = _games_split(ms)
                cost_ok = abs(a_cost - r["total"]) <= tolerance
                games_ok = (r["games"] is None) or (a_wc == r["games"])
                if cost_ok and games_ok:
                    alt = {**base, "TV Total Cost": a_cost,
                           "# Games w/Cost": a_wc, "# Games w/o Cost": a_woc}
                    not_reconciled.append({**alt, **variances(a_cost, a_wc, a_woc, alt_email),
                                           "Notes": "different email address", "_p": 2})
                    break
            else:
                not_reconciled.append({**base, **variances(0.0, 0, 0), "Notes": "Not bought in", "_p": 0})
        else:
            not_reconciled.append({**base, **variances(0.0, 0, 0), "Notes": "Not bought in", "_p": 0})

    reconciled.sort(key=lambda x: (x["Team"], x["Email"]))
    not_reconciled.sort(key=lambda x: (x["_p"], -x["HAL Total Cost"]))
    return reconciled, not_reconciled


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #

ARIAL = "Arial"
CUR = "$#,##0"
THIN = Side(style="thin", color="000000")
HAL_FILL = PatternFill("solid", fgColor=Color(theme=8, tint=0.7999816888943144))
TV_FILL = PatternFill("solid", fgColor=Color(theme=8, tint=0.0))
VAR_FILL = PatternFill("solid", fgColor="FFC000")
NAVY = "1F3864"
BLUE = "2E5496"
CENTER = Alignment(horizontal="center")

RECON_COLS = ["Team", "Email", "Full/Partial", "Section", "Row", "Seats", "Qty",
              "# Games", "Total Cost", "Total Cost", "# Games w/Cost", "# Games w/o Cost"]
RECON_SRC = ["Team", "Email", "Full/Partial", "Section", "Row", "Seats", "Qty",
             "# Games", "HAL Total Cost", "TV Total Cost", "# Games w/Cost", "# Games w/o Cost"]
RECON_W = [22.3, 46.3, 15.6, 12.4, 9.6, 10.6, 8.6, 13.4, 14.6, 13.0, 22.0, 13.0]
RECON_BANDS = [("per HAL", HAL_FILL, 1, 9), ("per TicketVault", TV_FILL, 10, 12)]
RECON_COST = {9, 10}

NR_COLS = ["Notes", "Team", "Email", "Full/Partial", "Section", "Row", "Seats", "Qty",
           "# Games", "Total Cost", "Total Cost", "# Games w/Cost", "# Games w/o Cost",
           "Total Cost", "Total Cost", "# Games w/Cost", "Email Address"]
NR_SRC = ["Notes", "Team", "Email", "Full/Partial", "Section", "Row", "Seats", "Qty",
          "# Games", "HAL Total Cost", "TV Total Cost", "# Games w/Cost", "# Games w/o Cost",
          "Var Total Cost", "Var Total Cost %", "Var # Games w/Cost", "Var Email Address"]
NR_W = [26.0, 20.0, 28.6, 15.6, 12.4, 9.6, 10.6, 8.6, 13.4, 14.6, 13.0, 20.1, 22.0,
        13.0, 13.0, 15.0, 30.0]
NR_BANDS = [("per HAL", HAL_FILL, 1, 10), ("per TicketVault", TV_FILL, 11, 13),
            ("Variances", VAR_FILL, 14, 17)]
NR_COST = {10, 11, 14}
NR_PCT = {15}


def _build_detail_tab(ws, headers, srcs, widths, rows, bands, cost_cols, pct_cols=frozenset()):
    ws.sheet_view.showGridLines = False
    for j, (w, h) in enumerate(zip(widths, headers), 1):
        ws.column_dimensions[get_column_letter(j)].width = max(w, len(str(h)) + 4)
    n = len(headers)
    for label, fill, first, last in bands:
        ws.merge_cells(start_row=1, start_column=first, end_row=1, end_column=last)
        c = ws.cell(1, first, label)
        c.font = Font(name=ARIAL, size=9, bold=True, color="000000")
        c.alignment = CENTER
        for cc in range(first, last + 1):
            ws.cell(1, cc).fill = fill
    for j, h in enumerate(headers, 1):
        cell = ws.cell(2, j, h)
        cell.font = Font(name=ARIAL, size=10, bold=True, color="000000")
        cell.alignment = CENTER
    for i, r in enumerate(rows, 3):
        for j, src in enumerate(srcs, 1):
            cell = ws.cell(i, j, r.get(src))
            cell.font = Font(name=ARIAL, size=9); cell.alignment = CENTER
            if j in cost_cols:
                cell.number_format = CUR
            elif j in pct_cols:
                cell.number_format = "0.00%"
    end = 2 + len(rows)
    lefts = {b[2] for b in bands}
    rights = {b[3] for b in bands}
    for rr in range(1, end + 1):
        for col in lefts:
            ws.cell(rr, col).border = Border(left=THIN, right=ws.cell(rr, col).border.right)
        for col in rights:
            ws.cell(rr, col).border = Border(left=ws.cell(rr, col).border.left, right=THIN)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}{end}"


def _src_val(v):
    if v is None or isinstance(v, (str, int, float, bool, dt.datetime, dt.date)):
        return v
    return str(v)


def _build_source_tab(ws, header, blocks):
    """Dump raw source rows (header + data). `blocks` is a list of (header, rows)
    so multiple files for one company stack. `header` is a fallback header."""
    ws.sheet_view.showGridLines = False
    max_cols = 1
    r = 1
    first = True
    for blk_header, blk_rows in blocks:
        hdr = blk_header or header
        for j, h in enumerate(hdr, 1):
            cell = ws.cell(r, j, _src_val(h))
            cell.font = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=NAVY); cell.alignment = CENTER
        max_cols = max(max_cols, len(hdr))
        hdr_row = r
        r += 1
        for row in blk_rows:
            for j, v in enumerate(row, 1):
                ws.cell(r, j, _src_val(v)).font = Font(name=ARIAL, size=9)
            r += 1
        if first:
            ws.freeze_panes = f"A{hdr_row + 1}"
            ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(max(max_cols, 1))}{r - 1}"
            first = False
        r += 1  # blank row between stacked files
    for j in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16


def build_workbook(company, league, year, as_of, reconciled, not_reconciled,
                   hal_total, tolerance, hal_blocks, pv_header, pv_rows):
    t_rec = [r for r in reconciled if not r.get("_parking")]
    p_rec = [r for r in reconciled if r.get("_parking")]
    t_nr = [r for r in not_reconciled if not r.get("_parking")]
    p_nr = [r for r in not_reconciled if r.get("_parking")]
    rec_n, nr_n = len(reconciled), len(not_reconciled)
    nbi = sum(1 for r in not_reconciled if r["Notes"] == "Not bought in")
    de = sum(1 for r in not_reconciled if r["Notes"] == "different email address")
    mismatch = nr_n - nbi - de

    wb = Workbook()

    # ---- Summary -------------------------------------------------------- #
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 72.7
    ws.column_dimensions["B"].width = 12.0
    ws.merge_cells("A1:B1")
    t = ws.cell(1, 1, "Seasons Review")
    t.font = Font(name=ARIAL, size=15, bold=True, color=NAVY); t.alignment = CENTER

    def info(r, text, size, bold):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(r, 1, text); c.font = Font(name=ARIAL, size=size, bold=bold, color=NAVY); c.alignment = CENTER

    info(3, f"Company:  {company}", 11, True)
    info(4, f"League:  {league}", 10, False)
    info(5, f"Year:  {year}", 10, False)
    info(6, f"As Of:  {as_of}", 10, False)

    def bar(r, text):
        c = ws.cell(r, 1, text); c.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE); c.alignment = CENTER
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=BLUE)

    def hd(r, a, b):
        for col, val in ((1, a), (2, b)):
            c = ws.cell(r, col, val); c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = CENTER

    def line(r, a, b, bold=False):
        ca = ws.cell(r, 1, a); cb = ws.cell(r, 2, b)
        for c in (ca, cb):
            c.font = Font(name=ARIAL, size=10, bold=bold); c.alignment = CENTER

    bar(8, "RESULT"); hd(9, "Metric", "Count")
    line(10, "Reconciled", len(t_rec))
    line(11, "Not Reconciled", len(t_nr))
    line(12, "Parking Reconciled", len(p_rec))
    line(13, "Parking Not Reconciled", len(p_nr))
    line(14, "Total # HAL Records", hal_total, bold=True)
    bar(16, "NOT RECONCILED — BY REASON"); hd(17, "Reason", "Count")
    line(18, "Not bought in", nbi)
    line(19, "Different email address", de)
    line(20, "Cost / games mismatch", mismatch)
    line(21, "TOTAL", nr_n, bold=True)

    # ---- Tickets: Reconciled / Not Reconciled --------------------------- #
    _build_detail_tab(wb.create_sheet("Reconciled"), RECON_COLS, RECON_SRC, RECON_W,
                      t_rec, RECON_BANDS, RECON_COST)
    _build_detail_tab(wb.create_sheet("Not Reconciled"), NR_COLS, NR_SRC, NR_W,
                      t_nr, NR_BANDS, NR_COST, NR_PCT)

    # ---- Parking: Reconciled / Not Reconciled --------------------------- #
    _build_detail_tab(wb.create_sheet("Parking Reconciled"), RECON_COLS, RECON_SRC, RECON_W,
                      p_rec, RECON_BANDS, RECON_COST)
    _build_detail_tab(wb.create_sheet("Parking Not Reconciled"), NR_COLS, NR_SRC, NR_W,
                      p_nr, NR_BANDS, NR_COST, NR_PCT)

    # ---- Source data ---------------------------------------------------- #
    _build_source_tab(wb.create_sheet("HAL"), hal_blocks[0][0] if hal_blocks else [], hal_blocks)
    _build_source_tab(wb.create_sheet("Purchase Details"), pv_header, [(pv_header, pv_rows)])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read(), {"reconciled": rec_n, "not_reconciled": nr_n,
                        "tickets_reconciled": len(t_rec), "tickets_not_reconciled": len(t_nr),
                        "parking_reconciled": len(p_rec), "parking_not_reconciled": len(p_nr),
                        "not_bought_in": nbi, "different_email": de,
                        "mismatch": mismatch, "clean": nr_n == 0}


# --------------------------------------------------------------------------- #
# Routes
# --------------------------------------------------------------------------- #

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/options")
def options():
    return jsonify({"leagues": LEAGUES, "years": YEARS, "companies": COMPANY_NAMES})


@app.route("/process", methods=["POST"])
def process():
    hal_files = [f for f in request.files.getlist("hal") if f.filename]
    hal_companies = request.form.getlist("hal_company")
    details_files = [f for f in request.files.getlist("details") if f.filename]
    league = (request.form.get("league") or "").strip()
    year = (request.form.get("year") or "").strip()
    as_of = (request.form.get("as_of") or "").strip()

    if league not in LEAGUES:
        return jsonify({"error": "Please choose a League."}), 400
    if year not in YEARS:
        return jsonify({"error": "Please choose a Year."}), 400
    if not as_of:
        return jsonify({"error": "Please choose an As Of Date."}), 400
    as_of_fmt = _fmt_asof(as_of)
    if not hal_files:
        return jsonify({"error": "Please upload at least one HAL season ticket database."}), 400
    if not details_files:
        return jsonify({"error": "Please upload at least one TicketVault Purchase Details file."}), 400

    raw_tol = (request.form.get("tolerance") or "").strip()
    tolerance = DEFAULT_TOLERANCE
    if raw_tol:
        try:
            tolerance = abs(float(raw_tol))
        except ValueError:
            return jsonify({"error": "Tolerance must be a number (dollars)."}), 400

    warnings = []
    try:
        # ---- HAL: records + raw source blocks, grouped by company ---------- #
        hal_by_company = OrderedDict()
        hal_src_by_company = {}
        excluded_total = 0
        for i, f in enumerate(hal_files):
            company = (hal_companies[i] if i < len(hal_companies) else "").strip()
            if not company:
                return jsonify({"error": f"Choose a company for “{f.filename}”."}), 400
            raw = _rows_from_upload(f.filename, f.read())
            recs, excluded = parse_hal(raw, f.filename, company, year, league)
            hal_by_company.setdefault(company, []).extend(recs)
            hidx, _ = _find_header_row(raw)
            hdr = raw[hidx] if hidx < len(raw) else []
            hal_src_by_company.setdefault(company, []).append((hdr, raw[hidx + 1:]))
            excluded_total += excluded
        if not any(hal_by_company.values()):
            return jsonify({"error": "No season-ticket records found in the HAL file(s)."}), 400
        if excluded_total:
            warnings.append(f"Excluded {excluded_total} non-active row(s) "
                            f"(deposits, waitlist, inquiries, cancellations) from the review.")

        # ---- Purchase Details: parse, scope by company -------------------- #
        pv_parsed, pv_header, detail_rows = [], [], 0
        for f in details_files:
            parsed, hdr, usable = parse_details(_rows_from_upload(f.filename, f.read()),
                                                f.filename, league)
            pv_parsed.extend(parsed)
            if hdr:
                pv_header = hdr
            detail_rows += usable
        if detail_rows == 0:
            return jsonify({"error": "No usable rows found in the Purchase Details file(s)."}), 400

        # build a company-scoped vault per broker (Master Mapping List). Vendor
        # rows are kept in the source dump but skipped when matching.
        vault = {}   # broker -> {primary, secondary, raw}
        for x in pv_parsed:
            broker = PVCOMPANY_TO_BROKER.get(x["company_norm"])
            if not broker:
                continue
            vb = vault.setdefault(broker, {"primary": defaultdict(list),
                                           "secondary": defaultdict(list), "raw": []})
            vb["raw"].append(x["raw"])
            if x["excluded_vendor"]:
                continue
            vb["primary"][(x["email"], x["team"])].append(x)
            if not x["is_parking"]:
                vb["secondary"][(x["team"], x["sec"], x["row"])].append(x)

        token = uuid.uuid4().hex
        folder = os.path.join(STORE_DIR, token)
        os.makedirs(folder, exist_ok=True)

        reports = []
        tot_rec = tot_nr = 0
        for company, recs in hal_by_company.items():
            vb = vault.get(company, {"primary": {}, "secondary": {}, "raw": []})
            if not vb["raw"]:
                mapped = BROKER_PVCOMPANIES.get(company)
                warnings.append(f"No Purchase Details found for {company}"
                                + (f" (expected TicketVault company: "
                                   f"{', '.join(sorted(mapped))})" if mapped else
                                   " — company not in the Master Mapping List")
                                + " — all its records will show as “Not bought in”.")
            reconciled, not_reconciled = reconcile(recs, vb["primary"], vb["secondary"], tolerance)
            data, m = build_workbook(company, league, year, as_of_fmt, reconciled, not_reconciled,
                                     len(recs), tolerance, hal_src_by_company.get(company, []),
                                     pv_header, vb["raw"])
            fname = (f"Seasons Review - {_safe_name(company)} - {_safe_name(league)} - "
                     f"{_safe_name(year)} - As Of {_safe_name(as_of_fmt)}.xlsx")
            with open(os.path.join(folder, fname), "wb") as fh:
                fh.write(data)
            tot_rec += m["reconciled"]
            tot_nr += m["not_reconciled"]
            reports.append({
                "company": company, "records": len(recs),
                "reconciled": m["reconciled"], "not_reconciled": m["not_reconciled"],
                "clean": m["clean"], "filename": fname,
                "download_url": f"/download/{token}/{fname}",
            })

        reports.sort(key=lambda x: x["company"].lower())

        if len(reports) > 1:
            zip_name = (f"Seasons Review - {_safe_name(league)} - {_safe_name(year)} - "
                        f"As Of {_safe_name(as_of_fmt)}.zip")
            with zipfile.ZipFile(os.path.join(folder, zip_name), "w",
                                 zipfile.ZIP_DEFLATED) as zf:
                for rep in reports:
                    zf.write(os.path.join(folder, rep["filename"]), rep["filename"])
            download_all_url = f"/download/{token}"
            download_all_name = zip_name
        else:
            download_all_url = reports[0]["download_url"]
            download_all_name = reports[0]["filename"]

        _cleanup_old()
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    total_records = sum(len(v) for v in hal_by_company.values())
    return jsonify({
        "league": league, "year": year, "as_of": as_of_fmt,
        "companies": len(reports),
        "total_records": total_records, "reconciled": tot_rec, "not_reconciled": tot_nr,
        "clean": tot_nr == 0,
        "reports": reports,
        "download_all_url": download_all_url, "download_all_name": download_all_name,
        "warnings": warnings,
    })


@app.route("/download/<token>")
def download_all(token):
    folder = os.path.join(STORE_DIR, os.path.basename(token))
    if not os.path.isdir(folder):
        abort(404)
    zips = [f for f in os.listdir(folder) if f.lower().endswith(".zip")]
    if zips:
        pick = zips[0]
        return send_file(os.path.join(folder, pick), mimetype="application/zip",
                         as_attachment=True, download_name=pick)
    xlsx = [f for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    if len(xlsx) == 1:
        pick = xlsx[0]
        return send_file(os.path.join(folder, pick),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=pick)
    abort(404)


@app.route("/download/<token>/<path:name>")
def download_one(token, name):
    folder = os.path.join(STORE_DIR, os.path.basename(token))
    safe = os.path.basename(name)
    path = os.path.join(folder, safe)
    if not os.path.isfile(path):
        abort(404)
    return send_file(path,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=safe)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
